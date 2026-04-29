import os
import uuid
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from models import get_cart_items

async def generate_order_excel(db_pool, user_id: int) -> str:
    """
    Генерирует Excel-файл со сметой на основе корзины пользователя из БД.
    Возвращает абсолютный путь к сгенерированному файлу.
    """
    # 1. Получаем корзину пользователя из БД
    cart_items = await get_cart_items(db_pool, user_id)

    if not cart_items:
        return ""

    # 2. Создаем Excel документ
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"

    # Базовые стили для красоты
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовки таблицы
    headers = ["№", "Артикул", "Наименование", "Цена за шт. (руб)", "Кол-во", "Сумма (руб)"]
    ws.append(headers)
    
    # Стилизуем шапку
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Заполняем таблицу данными
    total_sum = 0.0
    row_num = 2

    for index, item in enumerate(cart_items, start=1):
        qty = int(item["quantity"])
        price = float(item["price"])
        line_total = price * qty
        total_sum += line_total

        row_data = [
            index,
            item["articul"],
            item["name"],
            price,
            qty,
            line_total
        ]
        ws.append(row_data)

        # Стилизуем текущую строку
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            if col in [1, 4, 5, 6]: # Выравниваем числа по центру
                cell.alignment = center_align
                
        row_num += 1

    # 4. Добавляем итоговую строку
    ws.append(["", "", "", "", "ИТОГО:", total_sum])
    
    total_label_cell = ws.cell(row=row_num, column=5)
    total_value_cell = ws.cell(row=row_num, column=6)
    
    total_label_cell.font = bold_font
    total_value_cell.font = bold_font
    total_label_cell.alignment = center_align
    total_value_cell.alignment = center_align

    # 5. Настраиваем ширину колонок, чтобы текст не обрезался
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 15  # Артикул
    ws.column_dimensions['C'].width = 65  # Наименование (самая широкая)
    ws.column_dimensions['D'].width = 18  # Цена
    ws.column_dimensions['E'].width = 10  # Кол-во
    ws.column_dimensions['F'].width = 18  # Сумма

    # 6. Сохраняем во временный файл
    # Создаем директорию tmp, если ее вдруг нет
    tmp_dir = "/tmp/nobe_bot_orders"
    os.makedirs(tmp_dir, exist_ok=True)
    
    # Генерируем уникальное имя файла, чтобы избежать конфликтов при одновременных заказах
    filename = f"Смета_{datetime.now().strftime('%d-%m-%Y')}_{uuid.uuid4().hex[:6]}.xlsx"
    filepath = os.path.join(tmp_dir, filename)
    
    wb.save(filepath)
    
    return filepath